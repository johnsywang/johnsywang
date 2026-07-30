# -*- coding: utf-8 -*-
"""生成多Sheet Excel:类目汇总/Top20素材明细/标杆案例脚本/框架公式对照/形态映射"""
import csv, json, sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\johnsywang\Desktop\商品素材创意框架分析.xlsx")
data = json.load(open('out/report_data.json', encoding='utf-8'))

wb = openpyxl.Workbook()
HEAD = PatternFill("solid", fgColor="5B21B6")
HEADF = Font(color="FFFFFF", bold=True, size=11)
SUB = PatternFill("solid", fgColor="EDE9FE")
thin = Side(style="thin", color="D8D8E0")
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

def style_header(ws, ncol, row=1):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD; cell.font = HEADF; cell.alignment = CENTER; cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# Sheet1 形态汇总
ws = wb.active; ws.title = "1-形态汇总"
ws.append(["内容形态","消耗合计(万元)","覆盖类目数","Top20素材数","可下载素材数","加权平均CTR(%)","代表类目"])
style_header(ws, 7)
for f in data['forms']:
    ws.append([f['form'], f['cost_w'], f['ncat'], f['top20'], f['n'], f['ctr'], "、".join(f['cats'][:6])])
autosize(ws, [18,16,12,12,14,16,50])
for r in range(2, ws.max_row+1):
    ws.cell(r,7).alignment = WRAP
    for c in range(1,8): ws.cell(r,c).border = BORDER

# Sheet2 类目汇总
ws = wb.create_sheet("2-类目汇总")
ws.append(["消耗排名","类目","所属形态","可下载素材数","Top20实际数","Top20消耗合计(万元)","Top20平均CTR(%)"])
style_header(ws, 7)
for i, c in enumerate(data['cats'], 1):
    ws.append([i, c['cat'], c['form'], c['n'], c['top20'], c['cost_w'], c['ctr']])
autosize(ws, [10,26,18,14,12,20,18])
for r in range(2, ws.max_row+1):
    for cc in range(1,8): ws.cell(r,cc).border = BORDER

# Sheet3 Top20素材明细
ws = wb.create_sheet("3-Top20素材明细")
rows = list(csv.DictReader(open("out/material_top20.csv", encoding='utf-8-sig')))
ws.append(["类目","类目内排名","消耗(元)","CTR(%)","建议文件名","素材URL"])
style_header(ws, 6)
for r in rows:
    ws.append([r['类目'], int(r['类目内排名']), float(r['消耗(元)']), float(r['ctr(%)']), r['建议文件名'], r['素材URL']])
autosize(ws, [22,10,14,10,30,60])

# Sheet4 标杆案例脚本
ws = wb.create_sheet("4-标杆案例脚本")
ws.append(["类目","产品","消耗(万元)","CTR(%)","内容形态","分镜结构(画面)","口播脚本要点","创意手法"])
style_header(ws, 8)
cases = [
 ["饮料冲调","山药莲子百合粉",160.3,3.34,"A 食补滋补","①山林研发人员出镜(脖子冒汗吃山药) ②工厂叉车溯源(拒绝中间商) ③三原料木盘特写 ④白瓷碗冲泡演示","头顶冒汗吃茯苓脖子冒汗吃山药→成分堆叠(每100克625mg亚麻酸)→源头直供拒绝中间商→孝亲人群→周年庆拍一发六","反常识知识钩子+专家溯源+成分数据+孝亲促销"],
 ["保健滋补","羊参牛骨粉",49.5,7.75,"A 食补滋补","①紫衣女主播演播室伪新闻(腰椎晚期) ②中式屏风中老年口播 ③蓝帽子备案罐 ④帮扶计划促销","北京传疯了男子腰椎晚期→关节无力痛点→毛牛骨髓西洋参→45岁以上帮扶计划→每罐立减两位数","伪新闻权威+痛点恐吓+保健食品备案(CTR全场最高)"],
 ["传统滋补品","宁夏吊干枸杞",37.6,6.01,"A 食补滋补","①访谈室专家持麦(50岁女子把枸杞当零食) ②枸杞地产地实拍 ③大果特写无硫磺","悬念故事→树上自然风干无硫磺→源头厂家去中间商→中老年每天坚持→今天下单便宜100","悬念故事钩子+产地溯源+去中间商"],
 ["休闲食品","乳清蛋白威化棒",66.5,3.83,"C 零食食品","①老板办公室砍价剧情 ②产品盒展示 ③一大箱特写(拍一发四) ④试吃反馈","不爱吃黄瓜鸡蛋试试这个→乳清蛋白全麦科学配方→买两箱送两箱到手四大箱→老公赞不绝口","老板剧情+成分健康+极致量感促销+试吃口碑"],
 ["面部洗护","街头实测湿巾",101.3,3.81,"B 护肤美妆","①校园门口(震惊!不敢置信) ②医护街头帮学生洗手实测 ③效果展示","震惊体悬念→街头实测→医护信任背书→效果对比","剧情悬念+街头实测+医护背书"],
 ["彩妆","谷雨美颜膏",40.4,2.22,"B 护肤美妆","①素颜女性手持产品(不会化妆的) ②御龄抗皱奢润美颜膏特写 ③节日红色场景","不会化妆懒人痛点→抗皱修色卖点→夏天人手一瓶→品牌备案背书","素颜痛点+懒人友好+品牌备案"],
 ["口腔洗护","俊小白美白牙膏",45.9,1.71,"D 家清清洁","①牙结石猎奇特写(牙缝滂臭) ②研究员展示牙膏(不是普通美白牙膏) ③使用演示","猎奇痛点→介孔羟基磷灰石色修科技→差异化卖点→研究员背书","猎奇痛点特写+研究员差异化"],
 ["洗护清洁","锅具清洁剂",87.0,2.50,"D 家清清洁","①水槽脏锅特写(静置8小时视频加速中) ②戴手套刷洗剧情 ③去污对比","脏锅痛点→静置实测加速→去污效果对比→合规声明","脏污痛点+加速实测+前后对比"],
 ["头发洗护","米菲娜染发霜",46.2,3.00,"D 家清清洁","①农村田野素人抱娃(再也不用去理发店) ②街拍旗袍美女(看不见白发的女人)","省钱痛点→植物调理国妆特字→30-50分钟见效→效果向往","素人痛点+效果向往+街拍对比"],
 ["内衣裤袜","欧迪芬调整内衣",37.7,1.46,"E 服饰鞋包","①真人手持内衣(始于1980专柜高端) ②无钢圈细节展示 ③尺码表两件立减20常驻","露背装场景→无钢圈收副乳防下垂防外扩两层塑型纱线→睡午觉不用脱→尺码促销","卖点关键词密集轰炸+品牌+尺码促销"],
 ["生活电器","便携电器",33.8,2.71,"F 家电数码","①酒店行李箱(少带一半行李) ②衣柜找插座(四周没插座) ③功能演示","场景困境痛点→功能一步解决→便携参数→工厂直供促销","场景痛点共鸣+功能演示"],
 ["厨房用具","三禾厨具",37.2,4.47,"F 家电数码","①三禾智能工厂门口主播+老板同框 ②工厂溯源对接","工厂溯源→老板对接去中间商→品质背书→价格实惠","工厂溯源+老板砍价"],
 ["个护健康","C50冲牙器",20.0,1.74,"B 护肤美妆","①家人对话剧情(别拔了都臭了) ②高压水枪冲牙 ③科技原理演示","剧情演绎痛点→超声微气泡科技→清洁率99%→国际证书背书→优惠入手","剧情演绎+科技卖点+证书背书"],
]
for c in cases: ws.append(c)
autosize(ws, [16,18,12,10,14,44,48,32])
for r in range(2, ws.max_row+1):
    for cc in [6,7,8]: ws.cell(r,cc).alignment = WRAP
    ws.row_dimensions[r].height = 78

# Sheet5 框架公式对照
ws = wb.create_sheet("5-框架公式对照")
ws.append(["内容形态","创意框架公式","开场钩子手法","信任背书","卖点呈现","促销转化","合规要点"])
style_header(ws, 7)
formulas = [
 ["A 食补滋补冲调","反常识知识/病症钩子 × 专家溯源 × 成分数据可视化 × 去中间商 × 孝亲+周年庆促销","反常识知识点/病症恐吓/悬念故事","研发员出镜+蓝帽子/国食健注号+源头基地","原料特写+配方克数(每100克含X毫克)","拍N发M/周年庆半价/给父母试试/错过又等一年","视频仅呈现创意·无医疗暗示·非医疗工作人员"],
 ["B 护肤美妆个护","痛点/悬念/剧情钩子 × 古法配方或成分科技 × 效果特写 × 备案信任 × 懒人友好","震惊剧情/素颜痛点/猎奇实测","国妆特字备案号+研究员/创始人IP+网友反馈","配方调制(克数)/成分科技/上脸演示","夏天人手一瓶+备案背书","使用效果因人而异·非医疗"],
 ["C 零食生鲜食品","食欲特写钩子 × 老板砍价剧情 × 成分健康背书 × 极致量感促销 × 试吃口碑","食欲特写/老板剧情/如果只能吃一种","厂家直播间+七天无理由+销量口碑分","配料成分展示/独立包装/一大箱实拍","拍一发四/买N送N/到手X大箱","产品广告合规·适量食用"],
 ["D 家清口腔清洁","脏污痛点特写 × 加速实测/剧情 × 科技成分卖点 × 前后对比 × 研究员背书","脏污猎奇特写(牙结石/脏锅/白发)","研究员字幕+国妆特字/器械证+加速实测","科技原理(介孔羟基磷灰石/超声微气泡)+清洁率99%","升级款+限时优惠","视频加速中·效果仅供参考"],
 ["E 服饰鞋包家纺","场景指向钩子 × 上身展示 × 卖点密集轰炸 × 尺码促销 × 品牌年份背书","场景指向(露背装看过来)/身材痛点","品牌始于X年/专柜高端+尺码表+面料科技","面料工艺细节特写+卖点关键词高频堆叠","两件立减/尺码引导","无"],
 ["F 家电数码智能","场景痛点钩子 × 功能演示 × 参数/竞品对比 × 工厂溯源 × 限时促销","场景困境(没插座/行李多/厨房难题)","工厂实拍+老板同框+参数展示","功能实操演示+参数卖点+竞品对比","工厂直供去中间商+限时","无"],
 ["G 医药健身其他","痛点/知识钩子 × 使用演示 × 专业资质背书 × 精准人群 × 促销","病症痛点/反常识知识","械字号/资质证书/专家人设","使用方式演示+成分/资质","人群福利+限时","非医疗暗示·舒缓辅助替代疗效"],
]
for f in formulas: ws.append(f)
autosize(ws, [16,46,26,32,34,28,30])
for r in range(2, ws.max_row+1):
    for cc in range(2,8): ws.cell(r,cc).alignment = WRAP
    ws.row_dimensions[r].height = 92

wb.save(OUT)
print(f"Excel已生成: {OUT}")
print(f"Sheet: {wb.sheetnames}")
